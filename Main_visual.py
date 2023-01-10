from PyQt5.QtWidgets import (QLineEdit, QPushButton, QLabel, QWidget, QApplication, QFrame,
                             QFileDialog, QMessageBox, QVBoxLayout, QHBoxLayout, QStyle, QDialog)
from PyQt5.QtGui import QFont, QPixmap
from PyQt5.QtCore import QDateTime, Qt
import openpyxl
from openpyxl import load_workbook
import time
import datetime

# фронтэнд ↓↓↓
class Main(QWidget):
    def __init__(self):
        super(Main, self).__init__()
        self.setWindowTitle("Аналитика v1.0 • Авто-Альянс")
        self.setMinimumSize(700, 350)
        self.setWindowIcon(self.style().standardIcon(QStyle.SP_FileDialogStart))

        vbox = QVBoxLayout()
        hbox_line_1 = QHBoxLayout()
        hbox_line_2 = QHBoxLayout()
        hbox_line_3 = QHBoxLayout()
        hbox_line_4 = QHBoxLayout()
        hbox_line_5 = QHBoxLayout()
        hbox_ok_close = QHBoxLayout()
        hbox_button_1 = QHBoxLayout()
        vbox.addWidget(self.hello_label())
        vbox.addStretch(0)
        hbox_button_1.addWidget(self.b_clear())
        hbox_button_1.setAlignment(Qt.AlignRight)
        vbox.addLayout(hbox_button_1)
        vbox.addWidget(self.horizontal_line())
        hbox_line_1.addWidget(self.book_label())
        hbox_line_1.addWidget(self.book_line_edit())
        vbox.addLayout(hbox_line_1)
        hbox_line_2.addWidget(self.book_ost_label())
        hbox_line_2.addWidget(self.book_ost_line_edit())
        vbox.addLayout(hbox_line_2)
        hbox_line_3.addWidget(self.book_price_label())
        hbox_line_3.addWidget(self.book_price_line_edit())
        vbox.addLayout(hbox_line_3)
        hbox_line_4.addWidget(self.book_const_label())
        hbox_line_4.addWidget(self.book_const_line_edit())
        vbox.addLayout(hbox_line_4)
        vbox.addWidget(self.horizontal_line())
        hbox_line_5.addWidget(self.book_file_save_label())
        hbox_line_5.addWidget(self.book_file_save_line_edit())
        vbox.addLayout(hbox_line_5)
        vbox.setSpacing(10)
        vbox.addWidget(self.horizontal_line())
        hbox_ok_close.addWidget(self.b_create())
        hbox_ok_close.addWidget(self.b_close())
        vbox.addLayout(hbox_ok_close)

        self.setLayout(vbox)

    def b_clear(self):
        b_clear = QPushButton(self)
        b_clear.setMinimumSize(65, 27)
        b_clear.setFont(QFont('Calibri', 8))
        b_clear.setText('Очистить\n всё')
        b_clear.setStyleSheet("""
                                                   QPushButton:hover { background-color: rgba(139, 0, 0);
                                                   border-radius: 10px;
                                                   border-style: ridge;
                                                   border-color: dark;
                                                   border-width: 2px; }
                               QPushButton:!hover { background-color: white;
                                                   border-style: ridge;
                                                   border-width: 2px;
                                                   border-radius: 10px;
                                                   border-color: dark;   }
                               QPushButton:pressed { background-color: rgb(255, 0, 0);
                                                   border-radius: 13px;}
                                   """)
        b_clear.clicked.connect(self.clean_all)
        return b_clear

    def clean_all(self):
        self.base_line = [self.book_line_edit, self.book_price_line_edit, self.book_const_line_edit,
                               self.book_ost_line_edit, self.book_file_save_line_edit]
        for line_edit in self.base_line:
            if len(line_edit.text()) != "":
                line_edit.setText("")

    def horizontal_line(self):
        self.decor_line = QFrame()
        self.decor_line.setFrameShape(QFrame.HLine)
        self.decor_line.setFrameShadow(QFrame.Sunken)
        return self.decor_line

    def hello_label(self):
        hello_label = QLabel('Аналитика v1.0')
        hello_label.setFont(QFont('Century Gothic', 20))
        hello_label.setAlignment(Qt.AlignCenter)
        return hello_label

    def book_label(self):
        book_label = QLabel("Выберите файл, содержащий остатки из 1С:")
        book_label.setMinimumSize(300, 10)
        return book_label

    def book_line_edit(self):
        self.book_line_edit = QLineEdit(self)
        self.book_line_edit.setMinimumSize(30, 30)
        serch_book_icon = self.book_line_edit.addAction(self.style().standardIcon(QStyle.SP_FileDialogContentsView), QLineEdit.TrailingPosition)
        serch_book_icon.triggered.connect(self.browse_files_book)
        self.book_line_edit.returnPressed.connect(self.dialog_save_name)
        return self.book_line_edit

    def book_ost_label(self):
        book_ost_label = QLabel("Выберите файл, содержащий остатки из базы:")
        book_ost_label.setMinimumSize(300, 10)
        return book_ost_label

    def book_ost_line_edit(self):
        self.book_ost_line_edit = QLineEdit(self)
        self.book_ost_line_edit.setMinimumSize(30, 30)
        serch_book_ost_icon = self.book_ost_line_edit.addAction(self.style().standardIcon(QStyle.SP_FileDialogContentsView), QLineEdit.TrailingPosition)
        serch_book_ost_icon.triggered.connect(self.browse_files_book_ost)
        self.book_ost_line_edit.returnPressed.connect(self.dialog_save_name)
        return self.book_ost_line_edit

    def book_price_label(self):
        book_price = QLabel("Выберите файл, содержащий прайс:")
        book_price.setMinimumSize(300, 10)
        return book_price

    def book_price_line_edit(self):
        self.book_price_line_edit = QLineEdit(self)
        self.book_price_line_edit.setMinimumSize(30, 30)
        serch_book_price_icon = self.book_price_line_edit.addAction(self.style().standardIcon(QStyle.SP_FileDialogContentsView), QLineEdit.TrailingPosition)
        serch_book_price_icon.triggered.connect(self.browse_files_book_price)
        self.book_price_line_edit.returnPressed.connect(self.dialog_save_name)
        return self.book_price_line_edit

    def book_const_label(self):
        book_const = QLabel("Выберите файл, содержащий константу:")
        book_const.setMinimumSize(300, 10)
        return book_const

    def book_const_line_edit(self):
        self.book_const_line_edit = QLineEdit(self)
        self.book_const_line_edit.setMinimumSize(30, 30)
        serch_book_const_icon = self.book_const_line_edit.addAction(self.style().standardIcon(QStyle.SP_FileDialogContentsView), QLineEdit.TrailingPosition)
        serch_book_const_icon.triggered.connect(self.browse_files_book_const)
        self.book_const_line_edit.returnPressed.connect(self.dialog_save_name)
        return self.book_const_line_edit

    def book_file_save_label(self):
        book_file_save_label = QLabel("Выберите папку, для сохранения новых файлов:")
        book_file_save_label.setMinimumSize(300, 10)
        return book_file_save_label

    def book_file_save_line_edit(self):
        self.book_file_save_line_edit = QLineEdit(self)
        self.book_file_save_line_edit.setMinimumSize(30, 30)
        serch_book_ost_icon = self.book_file_save_line_edit.addAction(self.style().standardIcon(QStyle.SP_DialogSaveButton), QLineEdit.TrailingPosition)
        serch_book_ost_icon.triggered.connect(self.browse_files_book_file_save)
        self.book_file_save_line_edit.returnPressed.connect(self.dialog_save_name)
        return self.book_file_save_line_edit

    def browse_files_book(self):
        browse_files_book = QFileDialog.getOpenFileName(self, 'Выберите файл, содержащий остатки из 1С...', '', 'xlsx files (*.xlsx)')
        self.book_line_edit.setText(browse_files_book[0])

    def browse_files_book_ost(self):
        browse_files_book_ost = QFileDialog.getOpenFileName(self, 'Выберите файл, содержащий остатки из базы...', '', 'xlsx files (*.xlsx)')
        self.book_ost_line_edit.setText(browse_files_book_ost[0])

    def browse_files_book_price(self):
        browse_files_book_price = QFileDialog.getOpenFileName(self, 'Выберите файл, содержащий прайс...', '', 'xlsx files (*.xlsx)')
        self.book_price_line_edit.setText(browse_files_book_price[0])

    def browse_files_book_const(self):
        browse_files_book_const = QFileDialog.getOpenFileName(self, 'Выберите файл, содержащий константу... ', '', 'xlsx files (*.xlsx)')
        self.book_const_line_edit.setText(browse_files_book_const[0])

    def browse_files_book_file_save(self):
        browse_files_book_file_save = QFileDialog.getExistingDirectory(self, "Выбор папки для сохранения...")
        self.book_file_save_line_edit.setText(browse_files_book_file_save)

    def b_create(self):
        b_create = QPushButton("Выполнить обработку • Enter", self)
        b_create.setMinimumSize(10, 40)
        b_create.setFont(QFont('Century Gothic', 8, QFont.Normal))
        b_create.setShortcut('Enter')
        b_create.setToolTip('Нажмите, для запуска программы')
        b_create.setStyleSheet("""
                                                   QPushButton:hover { background-color: green;
                                                   border-radius: 10px;
                                                   border-style: ridge;
                                                   border-color: dark;
                                                   border-width: 2px; }
                               QPushButton:!hover { background-color: white;
                                                   border-style: ridge;
                                                   border-width: 2px;
                                                   border-radius: 10px;
                                                   border-color: dark;   }
                               QPushButton:pressed { background-color: rgb(0, 255, 0);
                                                   border-radius: 17px;}
                                   """)

        b_create.clicked.connect(self.start_check)
        return b_create

    def b_close(self):
        b_close = QPushButton("Закрыть • Esc", self)
        b_close.setMinimumSize(50, 40)
        b_close.setFont(QFont('Century Gothic', 8, QFont.Normal))
        b_close.setShortcut('Esc')
        b_close.setToolTip('Нажмите, чтобы выйти')
        b_close.setStyleSheet("""
                                           QPushButton:hover { background-color: rgba(139, 0, 0);
                                                               border-radius: 10px;
                                                               border-style: ridge;
                                                               border-color: dark;
                                                               border-width: 2px; }
                                           QPushButton:!hover { background-color: white;
                                                                border-style: ridge;
                                                                border-width: 2px;
                                                                border-radius: 10px;
                                                                border-color: dark;   }
                                           QPushButton:pressed { background-color: rgb(255, 0, 0);
                                                                 border-radius: 17px;}
                                       """)

        b_close.clicked.connect(self.start_close)
        return b_close

    def start_close(self):
        self.close()

    def start_check(self):
        self.base_line_edit = [self.book_line_edit, self.book_price_line_edit, self.book_const_line_edit,
                               self.book_ost_line_edit, self.book_file_save_line_edit]
        for line_edit in self.base_line_edit:
            if len(line_edit.text()) == 0:
                self.showMessageBox('Внимание!',
                                    '<center style=font-size:11pt><FONT FACE="Century Gothic"><b><u>Вы не заполнили поля!</center></u></b>'
                                    '<center style=font-size:7pt><FONT FACE="Century Gothic">(нажмите <b>ОК</b> и попробуйте снова)</center>')
                return

        self.dialog_save_name()

    def start_analytics(self):
        self.start_time = time.time()
        self.ostatok_base_44_44dealer_46alyans()
        self.ostatok_46_46dealer_46BOSCH()
        print(f'отработла за {int(time.time() - self.start_time)} секунд')
        self.showMessageBox_done()


    def showMessageBox(self, title, message):
        msgBox = QMessageBox()
        msgBox.setWindowTitle(title)
        msgBox.setWindowIcon(self.style().standardIcon(QStyle.SP_FileDialogStart))
        msgBox.setIcon(QMessageBox.Warning)
        msgBox.setDetailedText('Для корректной работы программы необходимо заполнить все поля, после чего нажать на кнопку "Выполнить обработку"')
        msgBox.setText(message)
        msgBox.setStandardButtons(QMessageBox.Ok)
        msgBox.exec_()

    def showMessageBox_done(self):
        msgBox = QMessageBox()
        msgBox.setWindowTitle("Обработка завершена!")
        msgBox.setWindowIcon(self.style().standardIcon(QStyle.SP_DialogApplyButton))
        msgBox.setIcon(QMessageBox.Warning)
        msgBox.setDetailedText('Работа обработчика успешно завершена!\n'
                               f'Время выполения запроса: {int(time.time() - self.start_time)} секунд(ы).')
        msgBox.setText('<center style=font-size:9pt><FONT FACE="Century Gothic"><b><u>Работа обработчика успешно завершена!</center></u></b>'
                       '<center style=font-size:7pt><FONT FACE="Century Gothic">(нажмите <b>ОК</b> для продолжения)</center>')
        msgBox.setStandardButtons(QMessageBox.Ok)
        if msgBox.exec_() == QMessageBox.Ok:
            print("Пока Братик")

    def image_and(self):
        label = QLabel(self)
        pixmap = QPixmap('line.png')
        label.setPixmap(pixmap)
        return label

    def displayTime(self):
        self.b_create().setText(QDateTime.currentDateTime().toString())
        self.b_create().adjustSize()

    def dialog_save_name(self):
        self.dlg = CustomDialog(self)
        self.dlg.exec_()

# фронтэнд ↑↑↑
# бэкэнд ↓↓↓

    def ostatok_base_44_44dealer_46alyans(self):
        book = load_workbook(self.book_line_edit.text())
        book_price = load_workbook(self.book_price_line_edit.text())
        book_const = load_workbook(self.book_const_line_edit.text())
        book_rez = openpyxl.Workbook()
        book_al = openpyxl.Workbook()
        book_44_dealer = openpyxl.Workbook()
        sheet = book.active
        sheet_price = book_price.active
        sheet_44_dealer = book_44_dealer.active
        sheet_rez = book_rez.active
        sheet_al = book_al.active
        sheet_const = book_const.active

        sheet_rez.append(['Наименование', 'Производитель', 'Номер', 'Остаток', 'Кратность', 'Цена'])
        sheet_al.append(['Наименование', 'Производитель', 'Номер', 'Остаток', 'Кратность', 'Цена'])
        sheet_44_dealer.append(['Наименование', 'Производитель', 'Номер', 'Остаток', 'Кратность', 'Цена'])
        dict_ost = {}
        for row in sheet.iter_rows(min_row=11, max_row=None):
            if row[2].value != 0 and row[1].value:
                dict_ost[str(row[1].value)] = row[2].value
        dict_price = {}
        for row in sheet_price.iter_rows(min_row=2, max_row=None):
            a = dict(
                number=str(row[0].value),
                name=row[1].value,
                brand=row[2].value,
                partia=row[3].value,
                price_opt=row[4].value,
                price_uch=row[5].value,
                price_rrp=row[6].value,
            )
            dict_price[str(row[0].value)] = a

        for k, v in dict_ost.items():  # 44 и 44дилер
            k = str(k)
            if dict_price.get(k) and dict_price[k]['brand'] != 'BOSCH':
                a = [dict_price[k]['name'],
                     dict_price[k]['brand'],
                     dict_price[k]['number'], v + 2,
                     dict_price[k]['partia'],
                     dict_price[k]['price_rrp']]
                b = [dict_price[k]['name'],
                     dict_price[k]['brand'],
                     dict_price[k]['number'], v,
                     dict_price[k]['partia'],
                     dict_price[k]['price_opt']]
                sheet_44_dealer.append(a)
                sheet_rez.append(b)
            elif dict_price.get(k.rjust(10, '0')) and dict_price[k.rjust(10, '0')]['brand'] == 'BOSCH' and v <= 500:
                k = k.rjust(10, '0')
                a = [dict_price[k]['name'],
                     dict_price[k]['brand'],
                     dict_price[k]['number'], v + 2,
                     dict_price[k]['partia'],
                     dict_price[k]['price_opt']]
                b = [dict_price[k]['name'],
                     dict_price[k]['brand'],
                     dict_price[k]['number'], v,
                     dict_price[k]['partia'],
                     dict_price[k]['price_opt']]
                sheet_44_dealer.append(a)
                sheet_rez.append(b)
            elif dict_price.get(k.rjust(10, '0')) and dict_price[k.rjust(10, '0')]['brand'] == 'BOSCH':
                k = k.rjust(10, '0')
                b = [dict_price[k]['name'],
                     dict_price[k]['brand'],
                     dict_price[k]['number'], 1000,
                     dict_price[k]['partia'],
                     dict_price[k]['price_opt']]
                sheet_44_dealer.append(b)
                sheet_rez.append(b)

        for row in range(1, sheet_const.max_row):  # добаква константы
            name = sheet_const[row][0].value
            brand = sheet_const[row][1].value
            art = str(sheet_const[row][2].value)
            ost = int(sheet_const[row][3].value)
            partia = int(sheet_const[row][4].value)
            price_opt = sheet_const[row][5].value
            a = [name, brand, art, ost, partia, price_opt]
            sheet_rez.append(a)
            sheet_44_dealer.append(a)

        for k, v in dict_ost.items():  # 46 alyans
            k = str(k)
            if dict_price.get(k) and dict_price[k]['brand'] == 'SWF' or dict_price.get(k) and dict_price[k][
                'brand'] == 'MOTUL' or dict_price.get(k) and dict_price[k]['brand'] == 'MANDO' or dict_price.get(k) and \
                    dict_price[k]['brand'] == 'BSG':
                b = [dict_price[k]['name'],
                     dict_price[k]['brand'],
                     dict_price[k]['number'], v,
                     dict_price[k]['partia'],
                     dict_price[k]['price_uch']]
                sheet_al.append(b)
            elif dict_price.get(k.rjust(10, '0')) and dict_price[k.rjust(10, '0')]['brand'] == 'BOSCH':
                k = k.rjust(10, '0')
                b = [dict_price[k]['name'],
                     dict_price[k]['brand'],
                     dict_price[k]['number'], v,
                     dict_price[k]['partia'],
                     dict_price[k]['price_uch']]
                sheet_al.append(b)

        book_rez.save(f'{self.book_file_save_line_edit.text()}/{self.line_edit_name_2.text()} {datetime.date.today().strftime("%d.%m.%y")}.xlsx')
        book_rez.save(f'{self.book_file_save_line_edit.text()}/{self.line_edit_name_7.text()}.xlsx')
        book_al.save(f'{self.book_file_save_line_edit.text()}/{self.line_edit_name_6.text()} {datetime.date.today().strftime("%d.%m.%y")}.xlsx')
        book_44_dealer.save(f'{self.book_file_save_line_edit.text()}/{self.line_edit_name_3.text()} {datetime.date.today().strftime("%d.%m.%y")}.xlsx')
        book.close()
        book_rez.close()
        book_al.close()
        book_price.close()
        book_44_dealer.close()

    def ostatok_46_46dealer_46BOSCH(self):
        book_ost = load_workbook(self.book_ost_line_edit.text())
        book_price = load_workbook(self.book_price_line_edit.text())
        book_46rez = openpyxl.Workbook()
        book_46_bosch_rez = openpyxl.Workbook()
        book_46_dealer_rez = openpyxl.Workbook()
        book_44_bosch_rez = openpyxl.Workbook()
        sheet_46rez = book_46rez.active
        sheet_44_bosch_rez = book_44_bosch_rez.active
        sheet_ost = book_ost.active
        sheet_price = book_price.active # доступ к листу с ценами
        sheet_46_dealer_rez = book_46_dealer_rez.active # доступ к листу с дилерами
        sheet_46_bosch_rez = book_46_bosch_rez.active # доступ к листу с резервами

        sheet_46rez.append(['Наименование', 'Производитель', 'Номер', 'Остаток', 'Кратность', 'Цена'])
        sheet_46_bosch_rez.append(['Наименование', 'Производитель', 'Номер', 'Остаток', 'Кратность', 'Цена'])
        sheet_46_dealer_rez.append(['Наименование', 'Производитель', 'Номер', 'Остаток', 'Кратность', 'Цена'])
        sheet_44_bosch_rez.append(['Наименование', 'Производитель', 'Номер', 'Остаток', 'Кратность', 'Цена'])

        dict_ost = {}
        for row in sheet_ost.iter_rows(min_row=2, max_row=None):
            a = dict(
                number=row[2].value,
                name=row[0].value,
                brand=row[1].value,
                partia=row[4].value,
                price_opt=row[5].value,
                ost=row[3].value
            )
            dict_ost[row[2].value] = a
        dict_price = {}
        for row in sheet_price.iter_rows(min_row=2, max_row=None):
            a = dict(
                number=row[0].value,
                name=row[1].value,
                brand=row[2].value,
                partia=row[3].value,
                price_opt=row[4].value,
                price_uch=row[5].value,
                price_rrp=row[6].value
            )
            dict_price[row[0].value] = a

        for v in dict_ost.values():  # 46 и 46бош и 46дилер
            if v['ost'] != 'Остаток':
                if v['brand'] != 'MOTUL':
                    d = [v['name'], v['brand'], v['number'], v['ost'], v['partia'], v['price_opt']]
                    sheet_44_bosch_rez.append(d)
                if v['brand'] == 'MOTUL':
                    a = [v['name'], v['brand'], v['number'], '36', v['partia'], v['price_opt']]
                    b = [v['name'], v['brand'], v['number'], '22', v['partia'], v['price_opt']]
                    c = [v['name'], v['brand'], v['number'], '24', v['partia'], v['price_opt']]
                    sheet_46rez.append(a)
                    sheet_46_bosch_rez.append(b)
                    sheet_46_dealer_rez.append(c)

                else:
                    if int(v['ost']) <= 8:
                        a = [v['name'], v['brand'], v['number'], '8', v['partia'], v['price_opt']]
                        b = [v['name'], v['brand'], v['number'], '6', v['partia'], v['price_opt']]
                        c = [v['name'], v['brand'], v['number'], '10', v['partia'], v['price_opt']]
                        sheet_46rez.append(a)
                        sheet_46_bosch_rez.append(b)
                        sheet_46_dealer_rez.append(c)
                    elif int(v['ost']) > 8 and int(v['ost']) <= 20:
                        a = [v['name'], v['brand'], v['number'], '20', v['partia'], v['price_opt']]
                        b = [v['name'], v['brand'], v['number'], '40', v['partia'], v['price_opt']]
                        c = [v['name'], v['brand'], v['number'], '30', v['partia'], v['price_opt']]
                        sheet_46rez.append(a)
                        sheet_46_bosch_rez.append(b)
                        sheet_46_dealer_rez.append(c)
                    elif int(v['ost']) > 20 and int(v['ost']) <= 50:
                        a = [v['name'], v['brand'], v['number'], '50', v['partia'], v['price_opt']]
                        b = [v['name'], v['brand'], v['number'], '40', v['partia'], v['price_opt']]
                        c = [v['name'], v['brand'], v['number'], '40', v['partia'], v['price_opt']]
                        sheet_46rez.append(a)
                        sheet_46_bosch_rez.append(b)
                        sheet_46_dealer_rez.append(c)
                    elif int(v['ost']) > 50 and int(v['ost']) <= 100:
                        a = [v['name'], v['brand'], v['number'], '100', v['partia'], v['price_opt']]
                        b = [v['name'], v['brand'], v['number'], '120', v['partia'], v['price_opt']]
                        c = [v['name'], v['brand'], v['number'], '90', v['partia'], v['price_opt']]
                        sheet_46rez.append(a)
                        sheet_46_bosch_rez.append(b)
                        sheet_46_dealer_rez.append(c)
                    elif int(v['ost']) > 100 and int(v['ost']) <= 500:
                        a = [v['name'], v['brand'], v['number'], '500', v['partia'], v['price_opt']]
                        b = [v['name'], v['brand'], v['number'], '400', v['partia'], v['price_opt']]
                        c = [v['name'], v['brand'], v['number'], '300', v['partia'], v['price_opt']]
                        sheet_46rez.append(a)
                        sheet_46_bosch_rez.append(b)
                        sheet_46_dealer_rez.append(c)
                    else:
                        a = [v['name'], v['brand'], v['number'], '1000', v['partia'], v['price_opt']]
                        b = [v['name'], v['brand'], v['number'], '1000', v['partia'], v['price_opt']]
                        c = [v['name'], v['brand'], v['number'], '1200', v['partia'], v['price_opt']]
                        sheet_46rez.append(a)
                        sheet_46_bosch_rez.append(b)
                        sheet_46_dealer_rez.append(c)

        book_46rez.save(f'{self.book_file_save_line_edit.text()}/{self.line_edit_name_8.text()} {datetime.date.today().strftime("%d.%m.%y")}.xlsx')
        book_46_bosch_rez.save(f'{self.book_file_save_line_edit.text()}/{self.line_edit_name_4.text()}.xlsx')
        book_46_dealer_rez.save(f'{self.book_file_save_line_edit.text()}/{self.line_edit_name_5.text()} {datetime.date.today().strftime("%d.%m.%y")}.xlsx')
        book_44_bosch_rez.save(f'{self.book_file_save_line_edit.text()}/{self.line_edit_name_1.text()}.xlsx')
        book_46rez.close()
        book_46_bosch_rez.close()
        book_46_dealer_rez.close()
        book_44_bosch_rez.close()

# бэкэнд ↑↑↑
# фронтэнд ↓↓↓
    def line_edit_name_1(self):
        self.line_edit_name_1 = QLineEdit("44_BOSCH")
        return self.line_edit_name_1

    def line_edit_name_2(self):
        self.line_edit_name_2 = QLineEdit("44_Остатки(1)")
        return self.line_edit_name_2

    def line_edit_name_3(self):
        self.line_edit_name_3 = QLineEdit("44дилер_Остатки")
        return self.line_edit_name_3

    def line_edit_name_4(self):
        self.line_edit_name_4 = QLineEdit("46_BOSCH")
        return self.line_edit_name_4

    def line_edit_name_5(self):
        self.line_edit_name_5 = QLineEdit("46дилер_Остатки")
        return self.line_edit_name_5

    def line_edit_name_6(self):
        self.line_edit_name_6 = QLineEdit("46учётАльянс_Остатки")
        return self.line_edit_name_6

    def line_edit_name_7(self):
        self.line_edit_name_7 = QLineEdit("ost_base")
        return self.line_edit_name_7

    def line_edit_name_8(self):
        self.line_edit_name_8 = QLineEdit("44_Остатки")
        return self.line_edit_name_8

class CustomDialog(QDialog):
    def __init__(self, Main):
        super(CustomDialog, self).__init__()

        self.Main = Main

        self.setWindowTitle("Сохранение...")
        self.setFixedSize(300,300)

        label_hello = QLabel()
        label_hello.setText('<center style=font-size:8.5pt><FONT FACE="Century Gothic">Задайте имена для сохранения новых файлов:</center>'
                            '<center style=font-size:7pt><FONT FACE="Century Gothic">(или оставьте стандартные и нажмите <b>ОК</b> для обработки)</center>')

        label_line_edit_name_1 = QLabel('<FONT FACE="Century Gothic" style=font-size:8.5pt>44_BOSCH:')
        label_line_edit_name_1.setMinimumSize(125, 1)
        label_line_edit_name_2 = QLabel('<FONT FACE="Century Gothic" style=font-size:8.5pt>44_Остатки(1):')
        label_line_edit_name_2.setMinimumSize(125, 1)
        label_line_edit_name_3 = QLabel('<FONT FACE="Century Gothic" style=font-size:8.5pt>44дилер_Остатки:')
        label_line_edit_name_3.setMinimumSize(125, 1)
        label_line_edit_name_4 = QLabel('<FONT FACE="Century Gothic" style=font-size:8.5pt>46_BOSCH:')
        label_line_edit_name_4.setMinimumSize(125, 1)
        label_line_edit_name_5 = QLabel('<FONT FACE="Century Gothic" style=font-size:8.5pt>46дилер_Остатки:')
        label_line_edit_name_5.setMinimumSize(125, 1)
        label_line_edit_name_6 = QLabel('<FONT FACE="Century Gothic" style=font-size:8.5pt>46учётАльянс_Остатки:')
        label_line_edit_name_6.setMinimumSize(125, 1)
        label_line_edit_name_7 = QLabel('<FONT FACE="Century Gothic" style=font-size:8.5pt>ost_base:')
        label_line_edit_name_7.setMinimumSize(125, 1)
        label_line_edit_name_8 = QLabel('<FONT FACE="Century Gothic" style=font-size:8.5pt>44_Остатки:')
        label_line_edit_name_8.setMinimumSize(125, 1)

        self.vBox = QVBoxLayout()
        self.vBox.addWidget(label_hello)
        self.hBox_1 = QHBoxLayout()
        self.hBox_2 = QHBoxLayout()
        self.hBox_3 = QHBoxLayout()
        self.hBox_4 = QHBoxLayout()
        self.hBox_5 = QHBoxLayout()
        self.hBox_6 = QHBoxLayout()
        self.hBox_7 = QHBoxLayout()
        self.hBox_8 = QHBoxLayout()
        self.hBox_9 = QHBoxLayout()
        self.hBox_1.addWidget(label_line_edit_name_1)
        self.hBox_1.addWidget(self.Main.line_edit_name_1())
        self.vBox.addLayout(self.hBox_1)
        self.hBox_2.addWidget(label_line_edit_name_2)
        self.hBox_2.addWidget(self.Main.line_edit_name_2())
        self.vBox.addLayout(self.hBox_2)
        self.hBox_3.addWidget(label_line_edit_name_3)
        self.hBox_3.addWidget(self.Main.line_edit_name_3())
        self.vBox.addLayout(self.hBox_3)
        self.hBox_4.addWidget(label_line_edit_name_4)
        self.hBox_4.addWidget(self.Main.line_edit_name_4())
        self.vBox.addLayout(self.hBox_4)
        self.hBox_5.addWidget(label_line_edit_name_5)
        self.hBox_5.addWidget(self.Main.line_edit_name_5())
        self.vBox.addLayout(self.hBox_5)
        self.hBox_6.addWidget(label_line_edit_name_6)
        self.hBox_6.addWidget(self.Main.line_edit_name_6())
        self.vBox.addLayout(self.hBox_6)
        self.hBox_7.addWidget(label_line_edit_name_7)
        self.hBox_7.addWidget(self.Main.line_edit_name_7())
        self.vBox.addLayout(self.hBox_7)
        self.hBox_8.addWidget(label_line_edit_name_8)
        self.hBox_8.addWidget(self.Main.line_edit_name_8())
        self.vBox.addLayout(self.hBox_8)
        self.hBox_9.addWidget(self.button_ok())
        self.hBox_9.addWidget(self.button_close())
        self.vBox.addLayout(self.hBox_9)
        self.setLayout(self.vBox)

    def button_ok(self):
        button_ok = QPushButton("ОК")
        button_ok.setMinimumSize(10, 40)
        button_ok.setShortcut('Enter')
        button_ok.setFont(QFont('Century Gothic', 8, QFont.Normal))
        button_ok.setStyleSheet("""
                                                   QPushButton:hover { background-color: green;
                                                   border-radius: 10px;
                                                   border-style: ridge;
                                                   border-color: dark;
                                                   border-width: 2px; }
                               QPushButton:!hover { background-color: white;
                                                   border-style: ridge;
                                                   border-width: 2px;
                                                   border-radius: 10px;
                                                   border-color: dark;   }
                               QPushButton:pressed { background-color: rgb(0, 255, 0);
                                                   border-radius: 17px;}
                                   """)
        button_ok.clicked.connect(self.ok_dialog)

        return button_ok
    def button_close(self):
        button_close = QPushButton("Отмена • ESC")
        button_close.setMinimumSize(50, 40)
        button_close.setShortcut('Esc')
        button_close.setFont(QFont('Century Gothic', 8, QFont.Normal))
        button_close.setStyleSheet("""
                                           QPushButton:hover { background-color: rgba(139, 0, 0);
                                                               border-radius: 10px;
                                                               border-style: ridge;
                                                               border-color: dark;
                                                               border-width: 2px; }
                                           QPushButton:!hover { background-color: white;
                                                                border-style: ridge;
                                                                border-width: 2px;
                                                                border-radius: 10px;
                                                                border-color: dark;   }
                                           QPushButton:pressed { background-color: rgb(255, 0, 0);
                                                                 border-radius: 17px;}
                                       """)
        button_close.clicked.connect(self.close_dialog)
        return button_close

    def ok_dialog(self):
        self.close_dialog()
        self.Main.start_analytics()
    def close_dialog(self):
        self.close()
# фронтэнд ↑↑↑

if __name__ == ('__main__'):
    import sys
    app = QApplication(sys.argv)
    w = Main()
    w.show()
    sys.exit(app.exec_())